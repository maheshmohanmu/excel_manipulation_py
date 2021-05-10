import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
working_sheet = inv_file["Sheet1"]

# changes from testbranch
#comment added from local repo
# print(working_sheet.max_row)

'''Logic for calculating & printing total value of each Supplier (inventory * cost), calculate total for each row
and save in a new file'''
supplier_tot_invent_val = {}

for each_row in range(2, working_sheet.max_row + 1):
    supp_name = working_sheet.cell(each_row, 4).value
    row_val = working_sheet.cell(each_row, 5)
    row_val.value = working_sheet.cell(each_row, 2).value * working_sheet.cell(each_row, 3).value
    if supp_name in supplier_tot_invent_val:
        supplier_tot_invent_val[supp_name] += working_sheet.cell(each_row, 2).value * working_sheet.cell(each_row, 3).value
    else:
        supplier_tot_invent_val[supp_name] = working_sheet.cell(each_row, 2).value * working_sheet.cell(each_row, 3).value

print(supplier_tot_invent_val)
inv_file.save("inventory_total.xlsx")
#x Logic for calculating and printing total number of products for each supplier
supplier_prod_count_set = {}

for prod_row in range(2, working_sheet.max_row + 1):
    supp_name = working_sheet.cell(prod_row, 4).value
    if supp_name in supplier_prod_count_set:
        supplier_prod_count_set[supp_name] += 1
    else:
        supplier_prod_count_set[supp_name] = 1

print(supplier_prod_count_set)

for prod_row in range(2, working_sheet.max_row + 1 ):
    if working_sheet.cell(prod_row, 2).value < 10:
        prod_num = working_sheet.cell(prod_row, 1).value
        print(f"Product {prod_num} has inventory less than 10")
